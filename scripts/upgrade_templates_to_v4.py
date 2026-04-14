"""
Upgrade MarginCOS v3.2 client data templates to v4.

Applies 8 coordinated changes to the four sector templates:
  1. Region dropdown globalized (7-region list)
  2. Currency-agnostic copy in DV errors and row 4 display labels
  3. Sheet protection enabled with password
  4. Formula cells get hidden=True (in addition to existing locked=True)
  5. Example rows replaced with neutral global placeholders
  6. Tier accent borders on row 4 column header cells
  7. Tier badge prefixes on row 4 display labels (ESS, PRO)
  8. Start Here sheet version stamp updated + change note added

Sector-aware: SKU trio (FMCG, Manufacturing, Retail) share structure;
Logistics branches separately.

Idempotent: running twice produces the same output.

Usage:
    python scripts/upgrade_templates_to_v4.py --sector FMCG
    python scripts/upgrade_templates_to_v4.py --all
"""
import argparse
import json
import os
import re
import zipfile
from pathlib import Path
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Protection, Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------- config ----------

PROTECTION_PASSWORD = "margincos2026"

GLOBAL_REGIONS = [
    "North America",
    "Latin America",
    "EMEA",
    "Sub-Saharan Africa",
    "Middle East & North Africa",
    "Asia Pacific",
    "Other",
]
GLOBAL_REGIONS_DV_FORMULA = '"' + ",".join(GLOBAL_REGIONS) + '"'

TIER_COLORS = {
    "ESS": "0D8F8F",  # teal
    "PRO": "D4A843",  # gold
}

# Tier mapping by row-3 snake_case key. Authoritative source from Wole's spec.
# Row-1 banner layout doesn't align perfectly with tier boundaries
# (e.g. business_unit is under PRICING banner but is an IDENTITY column for tier purposes).
TIER_BY_KEY_SKU = {
    # IDENTITY → none
    "sku_id": None, "sku_name": None, "category": None, "segment": None, "business_unit": None,
    # PRICING · P1 → ESS
    "rrp": "ESS", "cogs_per_unit": "ESS", "gross_margin_pct": "ESS",
    "price_elasticity": "ESS", "proposed_price_change_pct": "ESS",
    # VOLUME → ESS
    "monthly_volume_units": "ESS", "monthly_revenue": "ESS",
    # COST PASS-THROUGH · P2 → PRO
    "primary_channel": "PRO", "channel_revenue_split": "PRO",
    "cogs_inflation_rate": "PRO", "pass_through_rate": "PRO",
    # TRADE EXECUTION · P4 → PRO
    "distributor_margin_pct": "PRO", "promo_depth_pct": "PRO",
    "promo_lift_pct": "PRO", "wtp_premium_pct": "PRO",
    # METADATA → none
    "region": None, "active": None,
}

TIER_BY_KEY_LOGISTICS = {
    # IDENTITY → none
    "lane_id": None, "lane_name": None, "route_region": None, "cargo_type": None,
    "fleet_division": None, "active": None,
    # LANE RATE INTELLIGENCE + BACKHAUL → ESS
    "contracted_rate_ngn": "ESS", "fully_loaded_cost_ngn": "ESS", "distance_km": "ESS",
    "market_rate_ngn": "ESS", "rate_sensitivity": "ESS", "proposed_rate_change_pct": "ESS",
    "min_margin_floor_pct": "ESS", "rate_headroom_pct": "ESS",
    "return_lane_id": "ESS", "backhaul_rate_ngn": "ESS", "backhaul_cost_ngn": "ESS",
    "backhaul_recovery_pct": "ESS",
    # FUEL COST RECOVERY + CUSTOMER MIX + LOAD PROFITABILITY → PRO
    "fuel_cost_per_km": "PRO", "driver_cost_per_trip": "PRO",
    "maintenance_cost_per_trip": "PRO", "toll_levy_per_trip": "PRO",
    "cost_inflation_pct": "PRO", "pass_through_rate": "PRO",
    "prior_period_cost_ngn": "PRO", "fx_exposure_pct": "PRO",
    "truck_id": "PRO", "truck_type": "PRO", "contract_type": "PRO",
    "customer_name": "PRO", "customer_margin_pct": "PRO", "rebate_pct": "PRO",
    "payment_terms_days": "PRO", "fuel_surcharge_clause": "PRO",
    "monthly_trips": "PRO", "discount_depth_pct": "PRO", "volume_response_pct": "PRO",
    # METADATA → none
    "operating_region": None,
}

# Neutral-placeholder example data
EXAMPLES_FMCG = [
    # sku_id, sku_name, category, segment, business_unit, rrp, cogs_per_unit, (gross_margin_pct formula), price_elasticity, proposed_price_change_pct, monthly_volume_units, (monthly_revenue formula), primary_channel, channel_split, cogs_inflation_rate, pass_through_rate, distributor_margin_pct, promo_depth_pct, promo_lift_pct, wtp_premium_pct, region, active
    ["SKU-001", "Premium Pasta 500g",     "Noodles & Pasta",         "Premium",    "Dry Goods",    450,   260,  None, -1.2, 0.05, 45000, None, "Modern Trade",  "MT:60,OM:30,OT:10", 0.08, 0.75, 0.12, 0.15, 0.20, 0.08, "EMEA",                         "Y"],
    ["SKU-002", "Vegetable Oil 1L",        "Edible Oils",             "Mass Market","Liquid Foods", 1800,  1450, None, -0.8, 0.03, 82000, None, "Open Market",   "OM:70,MT:20,WS:10", 0.12, 0.60, 0.09, 0.10, 0.15, 0.05, "Sub-Saharan Africa",           "Y"],
    ["SKU-003", "Instant Coffee 100g",     "Tea & Coffee",            "Mid-Range",  "Beverages",    2200,  1320, None, -1.5, 0.04, 28000, None, "Modern Trade",  "MT:55,EC:25,HO:20", 0.06, 0.80, 0.14, 0.12, 0.18, 0.12, "Asia Pacific",                 "Y"],
    ["SKU-004", "Laundry Detergent 2kg",   "Laundry & Fabric Care",   "Mass Market","Home Care",    3200,  2050, None, -0.9, 0.02, 54000, None, "Wholesale",     "WS:50,MT:30,OM:20", 0.10, 0.65, 0.10, 0.08, 0.12, 0.06, "Latin America",                "Y"],
    ["SKU-005", "Toothpaste 150ml",        "Oral Care",               "Premium",    "Personal Care",850,   380,  None, -1.1, 0.06, 96000, None, "Modern Trade",  "MT:65,EC:20,OT:15", 0.05, 0.78, 0.15, 0.14, 0.22, 0.14, "North America",                "Y"],
]

EXAMPLES_MANUFACTURING = [
    ["SKU-001", "Galvanised Steel Sheet 1mm","Steel & Structural Products", "Mid-Range",  "Metals Division",   28000, 19500, None, -0.7, 0.04, 1200, None, "Direct",        "DI:70,WS:20,OT:10", 0.15, 0.55, 0.10, 0.05, 0.08, 0.04, "EMEA",                "Y"],
    ["SKU-002", "HDPE Plastic Pellets 25kg", "General Raw Material Inputs", "Mass Market","Polymers Division", 12500, 8700,  None, -1.0, 0.03, 3400, None, "Wholesale",     "WS:60,DI:30,OT:10", 0.09, 0.70, 0.08, 0.04, 0.06, 0.03, "Asia Pacific",        "Y"],
    ["SKU-003", "Portland Cement 50kg Bag",  "Cement & Concrete Products",  "Mass Market","Construction",      4200,  2600,  None, -0.6, 0.02, 18000,None, "Open Market",   "OM:55,WS:35,OT:10", 0.11, 0.68, 0.09, 0.06, 0.08, 0.02, "Sub-Saharan Africa",  "Y"],
    ["SKU-004", "Aluminium Roofing Sheet",   "Roofing & Insulation Materials","Premium","Metals Division",     48000, 34000, None, -0.8, 0.05, 850,  None, "Direct",        "DI:80,WS:20",       0.13, 0.62, 0.11, 0.03, 0.07, 0.07, "North America",       "Y"],
    ["SKU-005", "PVC Pipe 4in x 6m",         "Pipes & Fittings",            "Mid-Range",  "Polymers Division", 6800,  4300,  None, -1.1, 0.03, 5200, None, "Wholesale",     "WS:50,OM:30,DI:20", 0.08, 0.72, 0.10, 0.07, 0.10, 0.05, "Latin America",       "Y"],
]

EXAMPLES_RETAIL = [
    ["SKU-001", "Cotton T-Shirt M",             "Men's Clothing",        "Mass Market","Apparel",       3500,   1400,  None, -1.3, 0.05, 2400, None, "Modern Trade","MT:65,EC:25,OT:10", 0.07, 0.70, 0.14, 0.18, 0.25, 0.10, "EMEA",              "Y"],
    ["SKU-002", "Android Smartphone 6.5in",     "Mobiles & Smartphones", "Mid-Range",  "Electronics",   85000,  58000, None, -1.4, 0.04, 1800, None, "E-Commerce",  "EC:65,MT:25,OT:10", 0.05, 0.82, 0.18, 0.15, 0.25, 0.20, "Asia Pacific",      "Y"],
    ["SKU-003", "Running Shoes Size 42",        "Footwear",              "Premium",    "Footwear",      18500,  8200,  None, -0.9, 0.06, 820,  None, "Modern Trade","MT:60,EC:30,OT:10", 0.06, 0.75, 0.18, 0.15, 0.22, 0.22, "North America",     "Y"],
    ["SKU-004", "LED Television 43in",          "Large Appliances",      "Mid-Range",  "Electronics",   165000, 118000,None, -1.1, 0.04, 420,  None, "Modern Trade","MT:50,EC:40,OT:10", 0.06, 0.78, 0.15, 0.12, 0.18, 0.14, "Latin America",     "Y"],
    ["SKU-005", "Backpack 30L Daypack",         "Bags & Accessories",    "Mass Market","Accessories",   8900,   3900,  None, -1.1, 0.03, 1100, None, "Wholesale",   "WS:45,MT:35,EC:20", 0.09, 0.68, 0.15, 0.14, 0.20, 0.09, "Sub-Saharan Africa","Y"],
]

# Logistics: 38 columns
# Order follows Logistics col A..AL
EXAMPLES_LOGISTICS = [
    ["LANE-001","Lagos-Abuja",              "West Corridor",         "General Cargo",    "Heavy Haulage",       "Y", 1360000,1020000,760,1420000,-0.7,0.05,0.20,0.08, None, 720000, 580000, 0.81,140,260,120, 45, 0.11, 0.65, 950000, 0.15, "TRK-NG01", "40-Tonne",    "Dedicated",  "Continental Haulage",  0.15, 0.08, 60, "Yes",    160, 0.12, 0.15, "Sub-Saharan Africa"],
    ["LANE-002","Rotterdam-Hamburg",        "Cross-Border",          "Container (40ft)", "Long Haul",           "Y", 480000, 360000, 480, 510000, -0.8,0.04,0.12,0.14, None, 270000, 200000, 0.82, 85,140, 60, 22, 0.06, 0.75, 340000, 0.08, "TRK-EU02", "Trailer",     "Spot",       "Global Freight EU",    0.18, 0.06, 30, "Yes",    320, 0.08, 0.12, "EMEA"],
    ["LANE-003","Mumbai-Pune",              "West Corridor",         "Passengers",       "Passenger Transport", "Y", 180000, 128000, 150, 195000, -1.1,0.05,0.15,0.10, None, 0,      0,      0.00, 45, 95, 35, 12, 0.07, 0.72, 120000, 0.06, "TRK-AP03", "Luxury Bus",  "Contracted", "City Coach Lines",     0.22, 0.04, 30, "Partial",480, 0.09, 0.13, "Asia Pacific"],
    ["LANE-004","Los Angeles-Phoenix",      "Cross-Border",          "Cold Chain",       "Long Haul",           "Y", 620000, 485000, 580, 660000, -0.9,0.03,0.15,0.10, None, 350000, 280000, 0.80, 95,180, 80, 35, 0.05, 0.70, 460000, 0.04, "TRK-NA04", "Refrigerated","Spot",       "West Coast Logistics", 0.20, 0.04, 45, "Yes",    180, 0.10, 0.14, "North America"],
    ["LANE-005","São Paulo-Rio de Janeiro", "Coastal Route",         "Bulk / Loose",     "Mid Haul",            "Y", 520000, 410000, 430, 555000, -0.9,0.04,0.16,0.11, None, 300000, 240000, 0.80,105,170, 75, 28, 0.08, 0.68, 380000, 0.07, "TRK-LA05", "10-Tonne",    "Spot",       "LatAm Cargo",          0.17, 0.06, 45, "No",     210, 0.09, 0.13, "Latin America"],
]

# Start Here change note (appended; idempotent - checks for marker text)
V4_CHANGE_NOTE = (
    "v4 updates: globalised region dropdown (7 world regions), currency-agnostic "
    "labels and error messages, sheet protection enabled to prevent accidental edits, "
    "formula logic hidden, tier badges (ESS/PRO) added to column headers. "
    "Capacity: 500 data rows (rows 5-504)."
)

TEMPLATES_DIR = Path(__file__).resolve().parents[1] / "public" / "downloads"

SECTOR_FILES = {
    "FMCG":          "MarginCOS_Data_Template_v3.2_FMCG.xlsx",
    "Manufacturing": "MarginCOS_Data_Template_v3.2_Manufacturing.xlsx",
    "Retail":        "MarginCOS_Data_Template_v3.2_Retail.xlsx",
    "Logistics":     "MarginCOS_Data_Template_v3.2_Logistics.xlsx",
}

# ---------- helpers ----------

def tier_for_column(ws, col_idx, sector):
    """Return tier (ESS | PRO | None) for a column based on its row-3 snake_case key."""
    key = ws.cell(3, col_idx).value
    if not isinstance(key, str):
        return None
    key = key.strip()
    table = TIER_BY_KEY_SKU if sector != "Logistics" else TIER_BY_KEY_LOGISTICS
    return table.get(key)


def strip_existing_badge(label):
    """Remove ESS/PRO/SCALE/ENTERPRISE prefix if present (for idempotency)."""
    if not isinstance(label, str):
        return label
    for prefix in ("ESS  ", "PRO  ", "SCALE  ", "ENT  ", "ENTERPRISE  "):
        if label.startswith(prefix):
            return label[len(prefix):]
    return label


def replace_currency_in_label(label):
    """Replace ₦/xxx with currency/xxx in row 4 display labels."""
    if not isinstance(label, str):
        return label
    return label.replace("₦/unit", "currency/unit") \
                .replace("₦/trip", "currency/trip") \
                .replace("₦ ", "currency ") \
                .replace("Monthly Revenue ₦", "Monthly Revenue (currency)") \
                .replace("₦", "")


# ---------- change functions ----------

def change_1_globalize_region(wb, ws, sector, audit):
    """Globalize operating_region (Logistics) or region (SKU trio).
    Updates BOTH the inline DV on the data sheet AND the Reference_Lists tab."""
    ref = wb["Reference_Lists"]

    def safe_clear_and_write_regions(col_idx, start_row, clear_span=25):
        """Clear region values in a bounded range (skip merged cells), then write 7 new values."""
        from openpyxl.cell.cell import MergedCell
        before_vals = []
        for r in range(start_row, start_row + clear_span):
            cell = ref.cell(r, col_idx)
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None:
                before_vals.append(cell.value)
            cell.value = None
        for i, region in enumerate(GLOBAL_REGIONS):
            cell = ref.cell(start_row + i, col_idx)
            if isinstance(cell, MergedCell):
                raise RuntimeError(f"Region target cell {cell.coordinate} is merged")
            cell.value = region
        return before_vals

    if sector == "Logistics":
        # Reference_Lists layout: Operating Region in col F, starts row 2
        region_col_idx = 6
        region_start_row = 2
        before = safe_clear_and_write_regions(region_col_idx, region_start_row)

        # Update DV for operating_region (col AL) to new range
        for dv in ws.data_validations.dataValidation:
            if "AL5" in str(dv.sqref) and dv.type == "list":
                dv.formula1 = "=Reference_Lists!$F$2:$F$8"

        # Remove route_region (col C) DV — convert to free-text field
        remaining = []
        removed_route = False
        for dv in ws.data_validations.dataValidation:
            sqref = str(dv.sqref)
            if "C5:C" in sqref and dv.type == "list":
                removed_route = True
                continue
            remaining.append(dv)
        ws.data_validations.dataValidation = remaining

        audit["1_region"] = {
            "reference_lists_column": "F",
            "operating_region_before": before,
            "operating_region_after": GLOBAL_REGIONS,
            "route_region_dv_removed": removed_route,
        }

    else:
        # SKU trio: Reference_Lists layout: Region in col H, starts row 4
        region_col_idx = 8
        region_start_row = 4
        before_ref = safe_clear_and_write_regions(region_col_idx, region_start_row, clear_span=25)

        # Update inline DV on col V
        before_dv = None
        for dv in ws.data_validations.dataValidation:
            sqref = str(dv.sqref)
            if "V5:V1004" in sqref and dv.type == "list":
                before_dv = dv.formula1
                dv.formula1 = GLOBAL_REGIONS_DV_FORMULA
                dv.error = "Select a region from the dropdown list."

        audit["1_region"] = {
            "sku_data_col": "V",
            "inline_dv_before": before_dv,
            "inline_dv_after": GLOBAL_REGIONS_DV_FORMULA,
            "reference_lists_col_H_before": before_ref,
            "reference_lists_col_H_after": GLOBAL_REGIONS,
        }


def change_2_currency_agnostic(wb, ws, sector, audit):
    """Update DV error messages and row 4 labels to remove ₦."""
    changes = []

    # DV error messages
    for dv in ws.data_validations.dataValidation:
        if dv.error and "₦" in dv.error:
            before = dv.error
            dv.error = (
                "Numbers only - do not include currency symbols "
                "(e.g. ₦, $, €, £) or commas."
            )
            changes.append({"dv_ranges": str(dv.sqref), "before": before, "after": dv.error})

    # Row 4 labels containing ₦
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(4, c)
        if isinstance(cell.value, str) and "₦" in cell.value:
            before = cell.value
            cell.value = replace_currency_in_label(cell.value)
            changes.append({"cell": cell.coordinate, "before": before, "after": cell.value})

    audit["2_currency"] = changes


def change_3_sheet_protection(wb, audit):
    """Enable sheet protection on all sheets with the password.

    openpyxl boolean semantics (verified empirically):
      - False → XML "0" (user's desired block flags for structure)
      - True  → XML "1" (user's desired allow flags for sort/autoFilter)
    """
    protected = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        ws.protection.sheet = True
        ws.protection.password = PROTECTION_PASSWORD
        # Structure: blocked
        ws.protection.insertColumns = False
        ws.protection.deleteColumns = False
        ws.protection.insertRows = False
        ws.protection.deleteRows = False
        # Navigation aids: allowed
        ws.protection.sort = True
        ws.protection.autoFilter = True
        ws.protection.enable()
        protected.append(sn)
    audit["3_protection"] = {
        "sheets": protected,
        "password_set": PROTECTION_PASSWORD,
        "insertColumns": "blocked",
        "deleteColumns": "blocked",
        "insertRows": "blocked",
        "deleteRows": "blocked",
        "sort": "allowed",
        "autoFilter": "allowed",
    }


def unlock_input_cells(ws, sector, audit):
    """Unlock all non-formula data cells so users can enter data when sheet is protected.

    Formula cells retain locked=True (and hidden=True, from change_4).
    Non-formula data cells (rows 5..max) get locked=False.
    """
    # Detect formula columns by checking row 5 of each column
    formula_cols = set()
    for c in range(1, ws.max_column + 1):
        v5 = ws.cell(5, c).value
        if isinstance(v5, str) and v5.startswith("="):
            formula_cols.add(c)

    data_start = 5
    data_end = 1004 if sector != "Logistics" else 504
    start_col = 1 if sector == "Logistics" else 2
    end_col = ws.max_column

    unlocked_count = 0
    for r in range(data_start, data_end + 1):
        for c in range(start_col, end_col + 1):
            if c in formula_cols:
                continue
            cell = ws.cell(r, c)
            # Preserve any hidden flag if present; input cells should be locked=False, hidden=False
            cell.protection = Protection(locked=False, hidden=False)
            unlocked_count += 1

    audit["unlock_input_cells"] = {
        "formula_cols_kept_locked": [get_column_letter(c) for c in sorted(formula_cols)],
        "cells_unlocked": unlocked_count,
        "data_range": f"rows {data_start}-{data_end}, cols {get_column_letter(start_col)}-{get_column_letter(end_col)}",
    }


def change_4_hide_formulas(wb, audit):
    """Set hidden=True on all cells whose value is a formula (starts with '=')."""
    counts = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        cnt = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.protection = Protection(locked=True, hidden=True)
                    cnt += 1
        counts[sn] = cnt
    audit["4_hidden_formulas"] = counts


def change_5_replace_examples(ws, sector, audit):
    """Replace example rows (5-9) with neutral placeholders and clear 10+."""
    examples = {
        "FMCG": EXAMPLES_FMCG,
        "Manufacturing": EXAMPLES_MANUFACTURING,
        "Retail": EXAMPLES_RETAIL,
        "Logistics": EXAMPLES_LOGISTICS,
    }[sector]

    if sector == "Logistics":
        start_col = 1  # col A
        # Formula columns in Logistics? None identified on row 5. Safe to write all.
        # rate_headroom_pct, distance_km etc. — all non-formula.
    else:
        start_col = 2  # col B (col A is blank)

    # Preserve formula cells (they have value starting with =)
    formula_cols_preserved = []
    for c in range(start_col, ws.max_column + 1):
        v5 = ws.cell(5, c).value
        if isinstance(v5, str) and v5.startswith("="):
            formula_cols_preserved.append(get_column_letter(c))

    # Write 5 example rows
    for i, row_data in enumerate(examples):
        excel_row = 5 + i
        for j, val in enumerate(row_data):
            c = start_col + j
            # Skip formula columns - preserve the formula by copying from row 5
            v5 = ws.cell(5, c).value
            if val is None and isinstance(v5, str) and v5.startswith("="):
                # Capture the original formula pattern
                base_formula = v5
                # Adjust row references: replace digit 5 with excel_row
                # Naive but correct for simple =IFERROR(1-H5/G5,0) patterns
                new_formula = base_formula.replace("5", str(excel_row)) if excel_row != 5 else base_formula
                # Fix: only replace row 5 refs, not arbitrary 5's — safer to use relative
                # For idempotency, rewrite directly
                import re
                def repl(m):
                    col = m.group(1); row = int(m.group(2))
                    # Only adjust rows that equal 5 (the example row reference)
                    if row == 5:
                        return f"{col}{excel_row}"
                    return m.group(0)
                new_formula = re.sub(r"([A-Z]+)(\d+)", repl, base_formula)
                ws.cell(excel_row, c).value = new_formula
            else:
                ws.cell(excel_row, c).value = val

    # Clear rows 10+ (data but not formatting)
    cleared = 0
    for r in range(5 + len(examples), ws.max_row + 1):
        for c in range(start_col, ws.max_column + 1):
            if ws.cell(r, c).value is not None:
                # Don't touch formula cells (in case there are any seeded beyond row 9)
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.startswith("="):
                    continue
                ws.cell(r, c).value = None
                cleared += 1

    audit["5_examples"] = {
        "rows_written": len(examples),
        "formula_cols_preserved": formula_cols_preserved,
        "cells_cleared_beyond_examples": cleared,
    }


def change_6_tier_accents(ws, sector, audit):
    """Apply tier accent borders to row 4 column header cells."""
    accents = []

    start_col = 1 if sector == "Logistics" else 2
    end_col = ws.max_column

    for c in range(start_col, end_col + 1):
        tier = tier_for_column(ws, c, sector)
        cell = ws.cell(4, c)
        existing = cell.border
        if tier is None:
            # Clear any top border for all-tier columns (IDENTITY, METADATA)
            new_border = Border(
                left=existing.left,
                right=existing.right,
                top=Side(style=None),
                bottom=existing.bottom,
            )
            cell.border = new_border
            continue
        color = TIER_COLORS[tier]
        new_border = Border(
            left=existing.left,
            right=existing.right,
            top=Side(style="thick", color=color),
            bottom=existing.bottom,
        )
        cell.border = new_border
        accents.append({"col": get_column_letter(c), "tier": tier, "color": color})
    audit["6_tier_accents"] = accents


def change_7_tier_badges(ws, sector, audit):
    """Prefix row 4 display labels with tier badge (ESS / PRO)."""
    badges_applied = []

    start_col = 1 if sector == "Logistics" else 2
    end_col = ws.max_column

    for c in range(start_col, end_col + 1):
        tier = tier_for_column(ws, c, sector)
        cell = ws.cell(4, c)
        current = cell.value
        if current is None:
            continue
        stripped = strip_existing_badge(current)
        if tier is None:
            cell.value = stripped  # ensure no badge
        else:
            cell.value = f"{tier}  {stripped}"
            badges_applied.append({"col": get_column_letter(c), "tier": tier, "label": cell.value})
    audit["7_tier_badges"] = badges_applied


def change_8_start_here(wb, sector, audit):
    """Update Start Here: version stamp + change note."""
    sh = wb["Start Here"]
    changes = []

    # Find and update v3.2 → v4 in any cell
    for row in sh.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "v3.2" in cell.value:
                before = cell.value
                cell.value = cell.value.replace("v3.2", "v4")
                changes.append({"cell": cell.coordinate, "before": before, "after": cell.value})

    # Add/refresh change note at B5. Idempotent: overwrite only if content differs.
    b5_val = sh["B5"].value
    if b5_val != V4_CHANGE_NOTE:
        sh["B5"] = V4_CHANGE_NOTE
        sh["B5"].font = Font(name="Arial", size=10, italic=True, color="475569")
        sh["B5"].alignment = Alignment(wrap_text=True, vertical="top")
        sh.row_dimensions[5].height = 48
        changes.append({"cell": "B5", "before": b5_val, "after": V4_CHANGE_NOTE})

    audit["8_start_here"] = changes


def relax_format_flags_post_save(xlsx_path, audit):
    """
    Patch sheetProtection XML on all worksheets to relax three flags that openpyxl
    silently forces to "1" (blocked) on every save: formatCells, formatColumns, formatRows.

    MUST run AFTER wb.save() because openpyxl resets these flags each time the workbook
    is serialized. Operates directly on the xlsx zip file on disk.

    Idempotent: target_flags are forced to "0" whether they are currently "0", "1",
    or missing. Running twice produces identical XML.

    NOTE: function name is `relax_format_flags_post_save` to avoid collision with
    the existing `change_5_replace_examples` function. Semantically this is the
    post-save "change 5" step requested in the brief.
    """
    import shutil
    tmp_path = str(xlsx_path) + ".tmp"
    target_flags = ["formatCells", "formatColumns", "formatRows"]

    sheets_patched = []

    with zipfile.ZipFile(xlsx_path, "r") as zin:
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                data = zin.read(item)

                if item.startswith("xl/worksheets/sheet") and item.endswith(".xml"):
                    xml = data.decode("utf-8")

                    def patch_protection(match):
                        tag = match.group(0)
                        for flag in target_flags:
                            # Force to "0" whether currently "1", "0", or absent.
                            if re.search(rf'{flag}="\d"', tag):
                                tag = re.sub(rf'{flag}="\d"', f'{flag}="0"', tag)
                            else:
                                # Inject before closing "/>" (self-closing tag)
                                if tag.endswith("/>"):
                                    tag = tag[:-2] + f' {flag}="0"/>'
                                else:
                                    tag = tag[:-1] + f' {flag}="0">'
                        return tag

                    new_xml, n = re.subn(
                        r"<sheetProtection[^/]*/>",
                        patch_protection,
                        xml,
                    )
                    if n > 0:
                        sheets_patched.append(item)
                    data = new_xml.encode("utf-8")

                zout.writestr(item, data)

    shutil.move(tmp_path, xlsx_path)
    audit["9_post_save_flag_relaxation"] = {
        "sheets_patched": sheets_patched,
        "flags_forced_to_zero": target_flags,
    }


# ---------- driver ----------

def upgrade_template(sector):
    src = TEMPLATES_DIR / SECTOR_FILES[sector]
    dst = TEMPLATES_DIR / f"MarginCOS_Data_Template_v4_{sector}.xlsx"

    wb = load_workbook(src)
    data_sheet_name = "Lane Data" if sector == "Logistics" else "SKU Data"
    ws = wb[data_sheet_name]

    audit = {"sector": sector, "source": str(src.name), "output": str(dst.name)}

    change_1_globalize_region(wb, ws, sector, audit)
    change_2_currency_agnostic(wb, ws, sector, audit)
    change_5_replace_examples(ws, sector, audit)   # before hide/protect so formulas are still writable
    change_7_tier_badges(ws, sector, audit)
    change_6_tier_accents(ws, sector, audit)
    change_8_start_here(wb, sector, audit)
    change_4_hide_formulas(wb, audit)              # mark formulas hidden (locked=True, hidden=True)
    unlock_input_cells(ws, sector, audit)          # unlock non-formula data cells so users can type
    change_3_sheet_protection(wb, audit)           # last: enable sheet protection

    wb.save(dst)
    # Post-save: patch sheetProtection XML directly (openpyxl resets these on save)
    relax_format_flags_post_save(dst, audit)
    return audit


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--sector", choices=list(SECTOR_FILES.keys()))
    p.add_argument("--all", action="store_true")
    args = p.parse_args()

    targets = list(SECTOR_FILES.keys()) if args.all else [args.sector] if args.sector else []
    if not targets:
        p.error("Specify --sector or --all")

    audits = []
    for s in targets:
        audit = upgrade_template(s)
        audits.append(audit)
        print(f"[OK] {s} → {audit['output']}")

    audit_path = TEMPLATES_DIR.parent.parent / "scripts" / "template_v4_audit.json"
    audit_path.parent.mkdir(exist_ok=True)
    with open(audit_path, "w") as f:
        json.dump(audits, f, indent=2, default=str)
    print(f"\nAudit log: {audit_path}")


if __name__ == "__main__":
    main()
